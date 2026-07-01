import sys
import subprocess

# Verificar e instalar dependencias automaticamente
missing_deps = []
try:
    import pandas as pd  # type: ignore
except ImportError:
    missing_deps.append("pandas")

try:
    import openpyxl  # type: ignore
except ImportError:
    missing_deps.append("openpyxl")

try:
    from influxdb_client import InfluxDBClient  # type: ignore
except ImportError:
    missing_deps.append("influxdb-client")

if missing_deps:
    print(f"[INFO] No se encontraron las siguientes dependencias: {', '.join(missing_deps)}. Instalando automaticamente...")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing_deps)
        import pandas as pd  # type: ignore
        from influxdb_client import InfluxDBClient  # type: ignore
        print("[INFO] Dependencias instaladas con exito.\n")
    except Exception as e:
        print(f"[ERROR] No se pudieron instalar las dependencias de forma automatica: {e}")
        sys.exit(1)

import os
import re
import datetime
import shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore

# --- CONFIGURACIÓN DE INFLUXDB ---
INFLUX_URL = "https://us-east-1-1.aws.cloud2.influxdata.com"
INFLUX_TOKEN = "iDPm7LINRPeSd69ukOm9ZI7DjpgmmCZVU5RiMF7lHve6SznyMWL1j_InqbmKXVjSXsEZWd4Q0OXGniot4VouWg=="
INFLUX_ORG = "cx-radar"
INFLUX_BUCKET = "qoe_metrics"
PROBE_ID = "HUB_BTO"
RANGO_HISTORICO = "-90d"
# ---------------------------------

# Rutas locales basadas en la ubicacion del script
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_LOGS_DIR = os.path.join(BASE_DIR, "logs")

def escanear_logs_locales():
    """Busca en logs/ y subcarpetas todos los archivos qoe-probe-YYYY-MM-DD.log y extrae las fechas."""
    fechas_existentes = set()
    patron = re.compile(r"^qoe-probe-(\d{4}-\d{2}-\d{2})\.log$")
    
    if os.path.exists(LOCAL_LOGS_DIR):
        print(f"[INFO] Escaneando carpetas de logs localmente en: {LOCAL_LOGS_DIR}")
        for root, dirs, files in os.walk(LOCAL_LOGS_DIR):
            for file in files:
                match = patron.match(file)
                if match:
                    fecha_str = match.group(1)
                    fechas_existentes.add(fecha_str)
    else:
        print(f"[INFO] La carpeta local '{LOCAL_LOGS_DIR}' no existe aun, se creara al descargar.")
    
    return fechas_existentes

def obtener_fechas_influx(client, local_tz):
    """Consulta InfluxDB para obtener todos los días únicos en que HUB_BTO reportó datos."""
    query = f'''
    from(bucket: "{INFLUX_BUCKET}")
      |> range(start: {RANGO_HISTORICO})
      |> filter(fn: (r) => r["_measurement"] == "qoe_probe_run")
      |> filter(fn: (r) => r["probe_id"] == "{PROBE_ID}")
      |> keep(columns: ["_time"])
    '''
    query_api = client.query_api()
    tables = query_api.query(query)
    
    fechas = set()
    for table in tables:
        for record in table.records:
            utc_time = record["_time"]
            local_time = utc_time.astimezone(local_tz)
            fechas.add(local_time.strftime("%Y-%m-%d"))
    return fechas

def descargar_dia(client, fecha, local_tz):
    """Consulta InfluxDB y genera el archivo log reconstruido para un día específico."""
    fecha_dt = datetime.datetime.strptime(fecha, "%Y-%m-%d")
    inicio_local = datetime.datetime.combine(fecha_dt, datetime.time.min).replace(tzinfo=local_tz)
    fin_local = datetime.datetime.combine(fecha_dt, datetime.time.max).replace(tzinfo=local_tz)
    
    start_time_utc = inicio_local.astimezone(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    end_time_utc = fin_local.astimezone(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    query = f'''
    from(bucket: "{INFLUX_BUCKET}")
      |> range(start: {start_time_utc}, stop: {end_time_utc})
      |> filter(fn: (r) => r["_measurement"] == "qoe_real_metrics" or r["_measurement"] == "qoe_http_check" or r["_measurement"] == "qoe_probe_run")
      |> filter(fn: (r) => r["probe_id"] == "{PROBE_ID}")
      |> pivot(rowKey:["_time"], columnKey: ["_field"], valueColumn: "_value")
    '''

    query_api = client.query_api()
    tables = query_api.query(query)
    
    records = []
    for table in tables:
        for record in table.records:
            records.append(record.values)
            
    if not records:
        return False

    # Separar summaries (qoe_probe_run) de targets
    summaries = [r for r in records if r.get("_measurement") == "qoe_probe_run"]
    targets = [r for r in records if r.get("_measurement") in ("qoe_real_metrics", "qoe_http_check")]

    # Diccionario para agrupar líneas por combinación de router
    # key: (router_brand, router_model, connection_type)
    # value: lista de tuplas (timestamp, linea_de_texto)
    from collections import defaultdict
    lineas_por_router = defaultdict(list)

    # Formateador de latencia/jitter (remueve .0 si es entero para coincidir con la sonda)
    def fmt_lat_jit(val):
        if val is None:
            return "0"
        val = float(val)
        if val.is_integer():
            return str(int(val))
        return str(round(val, 2))

    # Reconstruir cada ciclo de la sonda (run)
    for s in summaries:
        utc_end = s["_time"]
        local_end = utc_end.astimezone(local_tz)
        duration_ms = s.get("run_duration_ms", 0.0)
        
        # Obtener router metadata específico de esta corrida
        r_brand = s.get("router_brand") or "unknown"
        r_model = s.get("router_model") or "unknown"
        r_conn = s.get("connection_type") or "unknown"
        
        # Calcular el inicio estimado del ciclo
        start_time_local = local_end - datetime.timedelta(milliseconds=duration_ms)
        start_time_rounded = start_time_local + datetime.timedelta(milliseconds=500)
        start_time_rounded = start_time_rounded.replace(microsecond=0)
        
        # Filtrar targets que pertenecen a este run específico (margen de 5 segs)
        run_targets = []
        for t in targets:
            if start_time_local - datetime.timedelta(seconds=5) <= t["_time"] <= utc_end + datetime.timedelta(seconds=5):
                run_targets.append(t)
                
        # Si la corrida no tenía tags, pero los targets sí (por robustez):
        if r_brand == "unknown" or r_brand is None:
            for t in run_targets:
                if t.get("router_brand") and t.get("router_brand") != "unknown":
                    r_brand = t.get("router_brand")
                    r_model = t.get("router_model")
                    r_conn = t.get("connection_type")
                    break
                    
        router_key = (r_brand, r_model, r_conn)
        
        # Deducir el jitter restando el tiempo de inicio del run y el del primer target
        jitter_seconds = 0
        if run_targets:
            run_targets.sort(key=lambda x: x["_time"])
            earliest_target = run_targets[0]
            target_dur_ms = earliest_target.get("run_duration_ms", 0.0)
            target_start_time_utc = earliest_target["_time"] - datetime.timedelta(milliseconds=target_dur_ms)
            target_start_time_local = target_start_time_utc.astimezone(local_tz)
            
            jitter_td = target_start_time_local - start_time_rounded
            jitter_seconds = int(round(jitter_td.total_seconds()))
            if jitter_seconds < 0:
                jitter_seconds = 0

        # Escribir cabeceras de inicio del run
        start_ts_str = start_time_rounded.strftime("%Y-%m-%dT%H:%M:%S")
        lineas_por_router[router_key].append((start_time_rounded, f"[{start_ts_str}] [INFO] Starting QoE probe using config C:\\Program Files\\CX-RADAR\\config\\probe catalog.json\n"))
        
        if jitter_seconds > 0:
            lineas_por_router[router_key].append((start_time_rounded, f"[{start_ts_str}] [INFO] Applying start jitter of {jitter_seconds} second(s) to reduce fixed-schedule burst patterns.\n"))
            token_time = start_time_rounded + datetime.timedelta(seconds=jitter_seconds)
            token_ts_str = token_time.strftime("%Y-%m-%dT%H:%M:%S")
            lineas_por_router[router_key].append((token_time, f"[{token_ts_str}] [INFO] Using InfluxDB token from .env file 'C:\\Program Files\\CX-RADAR\\.env'.\n"))
        else:
            lineas_por_router[router_key].append((start_time_rounded, f"[{start_ts_str}] [INFO] Using InfluxDB token from .env file 'C:\\Program Files\\CX-RADAR\\.env'.\n"))

        # Escribir targets de este run
        for t in run_targets:
            t_time_local = t["_time"].astimezone(local_tz)
            ts_str = t_time_local.strftime("%Y-%m-%dT%H:%M:%S")
            measurement = t.get("_measurement")
            service = t.get("service", "unknown")
            endpoint = t.get("endpoint_name", "unknown")
            available = t.get("available", True)
            available = False if str(available).lower() == "false" else bool(available)

            if measurement == "qoe_real_metrics":
                if available:
                    download = t.get("download_speed") or t.get("download_mbps") or 0.0
                    upload = t.get("upload_speed") or t.get("upload_mbps") or 0.0
                    latency = t.get("latency") or t.get("latency_ms") or 0.0
                    jitter = t.get("jitter") or t.get("jitter_ms") or 0.0
                    rpm = t.get("rpm_responsiveness") or 0.0
                    tool = t.get("tool") or "speedtest"
                    upload_tool = t.get("upload_tool") or "upload"
                    upload_err_class = t.get("upload_error_class")
                    upload_err_detail = t.get("upload_error_detail")
                    
                    loss = t.get("packet_loss_percentage") or 0.0
                    gw_lat = t.get("gateway_latency_ms") or 0.0
                    bb_dl = t.get("bufferbloat_download_ms") or 0.0
                    bb_ul = t.get("bufferbloat_upload_ms") or 0.0
                    
                    log_msg = f"Target {service}/{endpoint} completed with download {download} Mbps, upload {upload} Mbps via {upload_tool}, latency {fmt_lat_jit(latency)} ms, jitter {fmt_lat_jit(jitter)} ms, rpm {rpm}, tool {tool}"
                    log_msg += f", loss {loss}%, gateway_latency {gw_lat} ms, bufferbloat_dl {bb_dl} ms, bufferbloat_ul {bb_ul} ms"
                    if upload_err_class:
                        log_msg += f", upload_error_class {upload_err_class}: {upload_err_detail}"
                else:
                    err_class = t.get("error_class") or "unavailable"
                    err_detail = t.get("error_detail") or "No additional detail was provided."
                    if err_class == "probe_exception":
                        log_msg = f"Target {service}/{endpoint} failed: {err_detail}"
                    else:
                        log_msg = f"Target {service}/{endpoint} completed with error_class {err_class}: {err_detail}"
                
                lineas_por_router[router_key].append((t_time_local, f"[{ts_str}] [INFO] {log_msg}\n"))

            elif measurement == "qoe_http_check":
                err_class = t.get("error_class")
                err_detail = t.get("error_detail")
                
                if not available and err_class == "probe_exception":
                    log_msg = f"Target {service}/{endpoint} failed: {err_detail}"
                else:
                    http_status = int(t.get("http_status") or t.get("http_code") or 0)
                    curl_exit_code = int(t.get("curl_exit_code") or t.get("exit_code") or 0)
                    time_total = t.get("time_total_ms") or t.get("total_time_ms") or 0.0
                    log_msg = f"Target {service}/{endpoint} completed with HTTP {http_status}, curl exit {curl_exit_code}, total {time_total} ms"
                
                lineas_por_router[router_key].append((t_time_local, f"[{ts_str}] [INFO] {log_msg}\n"))

        # Escribir resúmenes de fin de ciclo
        success_count = int(s.get("success_count", 0))
        failure_count = int(s.get("failure_count", 0))
        total_count = success_count + failure_count
        end_ts_str = local_end.strftime("%Y-%m-%dT%H:%M:%S")
        
        lineas_por_router[router_key].append((local_end, f"[{end_ts_str}] [INFO] Wrote {total_count} measurement lines to InfluxDB.\n"))
        lineas_por_router[router_key].append((local_end, f"[{end_ts_str}] [INFO] QoE probe finished. Success={success_count}, Failure={failure_count}, Duration={duration_ms} ms\n"))

    # Escribir un archivo log independiente para cada router encontrado
    for (r_brand, r_model, r_conn), lineas in lineas_por_router.items():
        # Ordenar cronológicamente
        lineas.sort(key=lambda x: x[0])
        
        # Determinar directorio
        if r_brand != "unknown" and r_model != "unknown" and r_conn != "unknown" and r_brand is not None:
            target_dir = os.path.join(LOCAL_LOGS_DIR, r_brand, r_model, r_conn)
        else:
            target_dir = LOCAL_LOGS_DIR

        os.makedirs(target_dir, exist_ok=True)
        
        nombre_archivo = f"qoe-probe-{fecha}.log"
        local_path = os.path.join(target_dir, nombre_archivo)
        
        with open(local_path, "w", encoding="utf-8") as f:
            for _, linea in lineas:
                f.write(linea)
                
    return True

def ejecutar_descarga():
    """Descarga los logs faltantes desde InfluxDB antes de correr el reporte."""
    local_tz = datetime.datetime.now().astimezone().tzinfo
    hoy_str = datetime.date.today().strftime("%Y-%m-%d")
    
    print("\n--- INICIANDO DESCARGA AUTOMÁTICA DE LOGS DESDE INFLUXDB ---")
    fechas_locales = escanear_logs_locales()
    client = InfluxDBClient(url=INFLUX_URL, token=INFLUX_TOKEN, org=INFLUX_ORG)
    
    try:
        fechas_influx = obtener_fechas_influx(client, local_tz)
        fechas_faltantes = (fechas_influx - fechas_locales) - {hoy_str}
        
        if not fechas_faltantes:
            print("[INFO] Todos los logs históricos están al día. No se requiere descargar nada.")
            return

        print(f"[INFO] Se encontraron {len(fechas_faltantes)} día(s) faltante(s) por descargar.")
        for fecha in sorted(fechas_faltantes):
            exito = descargar_dia(client, fecha, local_tz)
            if not exito:
                print(f"  -> No se encontraron datos para reconstruir el día {fecha}.")
    except Exception as e:
        print(f"[ERROR] Error durante la descarga automática: {e}")
    finally:
        client.close()
    print("--- DESCARGA AUTOMÁTICA FINALIZADA ---\n")

def mapear_arbol_de_logs():
    """Escanea la carpeta 'logs' local si existe para clasificar archivos activos e históricos."""
    base_dir = LOCAL_LOGS_DIR if os.path.isdir(LOCAL_LOGS_DIR) else "."
    
    active_files = []
    completed_files = []

    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if file.endswith('.log') or file.endswith('.txt'):
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, start=base_dir)
                parts = rel_path.replace('\\', '/').split('/')
                
                # Estructura: logs/Marca/Modelo/Conexion/archivo.log
                if len(parts) >= 4:
                    completed_files.append({
                        "path": full_path,
                        "marca": parts[-4],
                        "modelo": parts[-3],
                        "conexion": parts[-2]
                    })
                else:
                    # Archivos sueltos en la raíz de logs
                    active_files.append(full_path)
                    
    return active_files, completed_files

def main():
    print("=== PROCESADOR INTELIGENTE DE LOGS CX-RADAR ===")
    
    # 1. Descargar logs faltantes antes de escanear la carpeta
    ejecutar_descarga()
    
    # 2. Escaneo automático de la carpeta logs local (ahora incluye los descargados)
    active_files, completed_files = mapear_arbol_de_logs()
    
    if not active_files and not completed_files:
        print("[ERROR] No se encontro la carpeta 'logs' o no contiene archivos (.log o .txt). Saliendo...")
        return

    # 3. Configurar metadatos de la prueba activa si existe
    router_marca_act = ""
    router_modelo_act = ""
    conexion_tipo_act = ""
    
    if active_files:
        print(f"\n[INFO] Se detectaron {len(active_files)} archivos de la PRUEBA ACTIVA actual en la raiz de logs.")
        print("-------------------------------------------------------------------------------------------------\n")
        router_marca_act = input("Ingrese la MARCA del router activo: ").strip()
        router_modelo_act = input("Ingrese el MODELO del router activo: ").strip()
        conexion_tipo_act = input("Ingrese el tipo de conexión activa: ").strip()
        
        # Mover los archivos a la estructura organizada si el usuario introdujo todos los datos
        if router_marca_act and router_modelo_act and conexion_tipo_act:
            target_dir = os.path.join(LOCAL_LOGS_DIR, router_marca_act, router_modelo_act, conexion_tipo_act)
            os.makedirs(target_dir, exist_ok=True)
            
            new_active_files = []
            for path in active_files:
                file_name = os.path.basename(path)
                new_path = os.path.join(target_dir, file_name)
                try:
                    shutil.move(path, new_path)
                    new_active_files.append(new_path)
                except Exception as e:
                    print(f"[WARNING] No se pudo mover {path} a {new_path}: {e}")
                    new_active_files.append(path)
            active_files = new_active_files
    
    if completed_files:
        print(f"\n[INFO] Se detectaron {len(completed_files)} archivos históricos en carpetas estructuradas.")

    print("\n=== SELECCIÓN DE FORMATO DE SALIDA ===")
    print("1. Excel (.xlsx)")
    print("2. CSV (.csv)")
    print("3. Texto (.txt - Separado por tabulaciones)")
    opcion_formato = input("Seleccione el formato deseado (1, 2 o 3): ").strip()

    # 4. Consolidar tareas de procesamiento
    tareas_procesamiento = []
    for path in active_files:
        tareas_procesamiento.append({
            "path": path, "marca": router_marca_act, "modelo": router_modelo_act, "conexion": conexion_tipo_act
        })
    for f in completed_files:
        tareas_procesamiento.append({
            "path": f["path"], "marca": f["marca"], "modelo": f["modelo"], "conexion": f["conexion"]
        })

    # Regex de extracción de CX-Radar
    regex_success = re.compile(
        r'^\[(?P<Timestamp>[^\]]+)\]\s+\[(?:INFO|ERROR)\]\s+Target\s+(?P<Target>[^\s]+)\s+completed\s+with\s+'
        r'download\s+(?P<Download>[0-9,.]+)\s+Mbps,\s+upload\s+(?P<Upload>[0-9,.]+)\s+Mbps\s+via\s+'
        r'(?P<Transport>[^,]+),\s+latency\s+(?P<Latency>[0-9,.]+)\s+ms,\s+jitter\s+(?P<Jitter>[0-9,.]+)\s+ms,\s+'
        r'rpm\s+(?P<RPM>[0-9,.]+),\s+tool\s+(?P<Tool>[^,\n]+)'
        r'(?:,\s+loss\s+(?P<Loss>[0-9,.]+)\s*%)?'
        r'(?:,\s+gateway_latency\s+(?P<GatewayLatency>[0-9,.]+)\s+ms)?'
        r'(?:,\s+bufferbloat_dl\s+(?P<BufferbloatDL>[0-9,.-]+)\s+ms)?'
        r'(?:,\s+bufferbloat_ul\s+(?P<BufferbloatUL>[0-9,.-]+)\s+ms)?'
        r'.*$'
    )
    regex_error = re.compile(
        r'^\[(?P<Timestamp>[^\]]+)\]\s+\[(?:INFO|ERROR)\]\s+Target\s+(?P<Target>[^\s]+)\s+(?:completed\s+with\s+error_class|failed:)\s+(?P<Error>.+)$'
    )

    extracted_data = []

    # 5. Ciclo de procesamiento unificado
    for tarea in tareas_procesamiento:
        log_path = tarea["path"]
        r_marca = tarea["marca"]
        r_modelo = tarea["modelo"]
        c_tipo = tarea["conexion"]
        
        if not os.path.exists(log_path):
            continue
            
        with open(log_path, 'r', encoding='utf-8') as file:
            for line in file:
                line_clean = line.strip()
                
                # Match Línea Exitosa
                match_success = regex_success.match(line_clean)
                if match_success:
                    data = match_success.groupdict()
                    target_bruto = data['Target'].lower()
                    if 'netflix' in target_bruto: target_limpio = 'Netflix'
                    elif 'youtube' in target_bruto: target_limpio = 'Youtube'
                    elif 'amazon' in target_bruto or 'aws' in target_bruto: target_limpio = 'Amazon'
                    elif 'disney' in target_bruto: target_limpio = 'Disney+'
                    elif 'microsoft' in target_bruto: target_limpio = 'Microsoft'
                    else: target_limpio = data['Target']

                    fecha_limpia = data['Timestamp'].replace('T', ' ')
                    download = float(data['Download'].replace(',', '.'))
                    upload = float(data['Upload'].replace(',', '.'))
                    latency = float(data['Latency'].replace(',', '.'))
                    jitter = float(data['Jitter'].replace(',', '.'))
                    rpm = float(data['RPM'].replace(',', '.'))

                    loss = float(data.get('Loss', '0').replace(',', '.')) if data.get('Loss') else 0.0
                    gw_lat = float(data.get('GatewayLatency', '0').replace(',', '.')) if data.get('GatewayLatency') else 0.0
                    bb_dl = float(data.get('BufferbloatDL', '0').replace(',', '.')) if data.get('BufferbloatDL') else 0.0
                    bb_ul = float(data.get('BufferbloatUL', '0').replace(',', '.')) if data.get('BufferbloatUL') else 0.0

                    # Evaluaciones de Experiencia (CX)
                    alertas_ciclo = []
                    if latency > 100.0:
                        alertas_ciclo.append(f"Latencia crítica en {target_limpio}: Degradación severa del enrutamiento o saturación de enlaces internacionales.")
                    elif latency >= 60.0:
                        alertas_ciclo.append(f"Latencia en Alerta en {target_limpio}: Degradación perceptible en interactividad en tiempo real.")
                    if 0 < download <= 5.0:
                        alertas_ciclo.append(f"Caída crítica de velocidad en {target_limpio}: Outlier técnico severo. Rompe el piso de la experiencia (QoE), forzando buffering.")
                    if jitter >= 40.0:
                        alertas_ciclo.append(f"Pico masivo de Jitter en {target_limpio}: Inestabilidad severa en la cola de paquetes de la red local o nodo de peering.")
                    elif jitter > 30.0:
                        alertas_ciclo.append(f"Jitter Crítico en {target_limpio}: Elevada variación en entrega de paquetes. Causa audio robotizado o artefactos visuales.")
                    if jitter >= 20.0 and latency >= 50.0 and download > 5.0:
                        alertas_ciclo.append(f"Inestabilidad combinada en {target_limpio}: Degradación paralela de Jitter y latencia. Evidencia congestión transitoria.")
                    if 0 < rpm <= 1050.0:
                        alertas_ciclo.append(f"Alerta latente de procesamiento en {target_limpio}: Caída notable del RPM por estrés. Hardware o tablas del router operando cerca de su límite.")
                    if 0 < upload <= 55.0:
                        alertas_ciclo.append(f"Contracción aguda de subida en {target_limpio}: Caída severa en la velocidad de carga bajo escenarios de estrés, indicador clave de upstream bufferbloat.")

                    comentario_final = "\n".join(alertas_ciclo) if alertas_ciclo else ""

                    record = {
                        "Fecha_Hora": fecha_limpia, "Router_Marca": r_marca, "Router_Modelo": r_modelo, "Tipo_Conexion": c_tipo,
                        "Target": target_limpio, "Download_Mbps": download, "Upload_Mbps": upload, "Latency_ms": latency,
                        "Jitter_ms": jitter, "RPM": rpm,
                        "Loss_Pct": loss, "Gateway_Lat": gw_lat, "BB_DL": bb_dl, "BB_UL": bb_ul,
                        "Estado": "OK", "Error_Msg": "", "Comentarios": comentario_final
                    }
                    extracted_data.append(record)
                    continue

                # Match Línea con Error
                match_error = regex_error.match(line_clean)
                if match_error:
                    data = match_error.groupdict()
                    target_bruto = data['Target'].lower()
                    if 'netflix' in target_bruto: target_limpio = 'Netflix'
                    elif 'youtube' in target_bruto: target_limpio = 'Youtube'
                    elif 'amazon' in target_bruto or 'aws' in target_bruto: target_limpio = 'Amazon'
                    elif 'disney' in target_bruto: target_limpio = 'Disney+'
                    elif 'microsoft' in target_bruto: target_limpio = 'Microsoft'
                    else: target_limpio = data['Target']

                    fecha_limpia = data['Timestamp'].replace('T', ' ')
                    error_raw = data['Error'].strip()
                    error_presentado = error_raw.split(';')[0].strip()

                    comentario = ""
                    es_wifi = c_tipo and 'wifi' in c_tipo.lower()

                    if target_limpio == 'Youtube' and not es_wifi:
                        comentario = (
                            f"Error presentado: {error_presentado}\n"
                            f"Este error es devuelto por los servidores de YouTube. Suele ocurrir porque YouTube detecta las "
                            f"peticiones automatizadas como tráfico de bots y bloquea temporalmente la IP."
                        )
                    elif target_limpio == 'Amazon' and not es_wifi:
                        comentario = (
                            f"Error presentado: {error_presentado}\n"
                            f"Se trata de un timeout de red a nivel de socket exclusivo al descargar el asset de AWS. Esto apunta a una "
                            f"degradación de la ruta específica hacia el CDN de Amazon, pérdida severa de paquetes solo en esa conexión TCP."
                        )
                    else:
                        comentario = f"Error presentado: {error_presentado}"

                    if es_wifi:
                        comentario += (
                            "\nNota: Conexión inalámbrica (Wi-Fi) activa durante el fallo. "
                            "El error coincide con inestabilidad en el enlace de radio, pérdida de paquetes o desconexión temporal."
                        )

                    record = {
                        "Fecha_Hora": fecha_limpia, "Router_Marca": r_marca, "Router_Modelo": r_modelo, "Tipo_Conexion": c_tipo,
                        "Target": target_limpio, "Download_Mbps": 0.0, "Upload_Mbps": 0.0, "Latency_ms": 0.0,
                        "Jitter_ms": 0.0, "RPM": 0.0,
                        "Loss_Pct": 0.0, "Gateway_Lat": 0.0, "BB_DL": 0.0, "BB_UL": 0.0,
                        "Estado": "ERROR", "Error_Msg": error_raw, "Comentarios": comentario
                    }
                    extracted_data.append(record)

    if extracted_data:
        df = pd.DataFrame(extracted_data)
        df = df.sort_values(by=["Router_Marca", "Router_Modelo", "Fecha_Hora"]).reset_index(drop=True)
        df = df.drop(columns=['Estado', 'Error_Msg'], errors='ignore')

        df = df.rename(columns={
            "Fecha_Hora": "Fecha Hora", "Router_Marca": "Router Marca", "Router_Modelo": "Router Modelo",
            "Tipo_Conexion": "Tipo Conexión", "Target": "Target", "Download_Mbps": "Download (Mbps)",
            "Upload_Mbps": "Upload (Mbps)", "Latency_ms": "Latency (ms)", "Jitter_ms": "Jitter (ms)",
            "RPM": "RPM",
            "Loss_Pct": "Pérdida Paquetes (%)",
            "Gateway_Lat": "Latencia Gateway (ms)",
            "BB_DL": "Bufferbloat Descarga (ms)",
            "BB_UL": "Bufferbloat Subida (ms)",
            "Comentarios": "Comentarios"
        })

        base_filename = "Metricas_QoE_Routers"
        
        # FIJACIÓN DE RUTA CRUCIAL: Fuerza a guardar donde está físicamente el archivo cx__report.py / ejecutar.bat
        output_dir = os.path.dirname(os.path.abspath(__file__))

        # Exportaciones fijadas en la raíz del script
        if opcion_formato == "2":
            output_path = os.path.join(output_dir, f"{base_filename}.csv")
            df.to_csv(output_path, index=False, sep=';', encoding='utf-8-sig')
            print(f"\n¡ÉXITO! Archivo CSV Maestro generado en la raiz: {output_path}")
        elif opcion_formato == "3":
            output_path = os.path.join(output_dir, f"{base_filename}.txt")
            df.to_csv(output_path, index=False, sep='\t', encoding='utf-8')
            print(f"\n¡ÉXITO! Archivo de Texto Maestro generado en la raiz: {output_path}")
        else:
            output_path = os.path.join(output_dir, f"{base_filename}.xlsx")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name="Métricas QoE", index=False)
                workbook = writer.book
                worksheet = writer.sheets["Métricas QoE"]
                
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                data_font = Font(name="Calibri", size=11)
                thin_side = Side(border_style="thin", color="D9D9D9")
                cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                
                worksheet.row_dimensions[1].height = 26
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = cell_border
                    
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.font = data_font
                        cell.border = cell_border
                        col_letter = cell.column_letter
                        if col_letter == 'A':
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif col_letter in ['B', 'C', 'D', 'E']:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif col_letter in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '0.00'
                        elif col_letter == 'O':
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                for col in worksheet.columns:
                    col_letter = col[0].column_letter
                    if col_letter == 'O':
                        worksheet.column_dimensions[col_letter].width = 75
                    elif col_letter == 'A':
                        worksheet.column_dimensions[col_letter].width = 20
                    else:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

            print(f"\n¡ÉXITO! Archivo Excel Maestro (.xlsx) generado en la raiz: {output_path}")
    else:
        print("\n[AVISO] No se encontraron archivos de log validos dentro de la carpeta 'logs'.")

if __name__ == "__main__":
    main()